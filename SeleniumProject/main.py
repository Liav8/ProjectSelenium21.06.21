from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from usableFunctionalities import *
from Product import *
from CleanCart import *
import time
from openpyxl import *


class MainP:
    def __init__(self, driver):
        self.driver = driver

    def login(self, driver):
        wait = WebDriverWait(driver, 10)
        # wait.until(EC.element_to_be_clickable((By.ID, "menuUser")))
        login_page = driver.find_element_by_id("menuUser")
        login_page.click()
        # wait.until(EC.element_to_be_clickable((By.ID, "username")))
        time.sleep(1.5)
        driver.find_element_by_name("username").send_keys("Sa1234")
        driver.find_element_by_name("password").send_keys("Sa1234")
        driver.find_element_by_id("sign_in_btnundefined").click()

    def logout(self, driver):
        driver.find_element_by_css_selector('[id="menuUserLink"]').click()
        time.sleep(0.2)
        driver.find_element_by_css_selector('[id="menuUserLink"] [translate="Sign_out"]').click()
        time.sleep(0.2)

    def clearCart(self, driver):
        driver.find_element_by_xpath('//*[@id="menuCart"]').click()
        time.sleep(1)
        remove = driver.find_elements_by_css_selector('[class="fixedTableEdgeCompatibility"] [translate="REMOVE"]')
        if remove != []:
            for i in range(len(remove) - 1, -1, -1):
                remove[i].click()
                time.sleep(0.5)
        time.sleep(0.5)
        driver.find_element_by_css_selector('[class="logo"] [href="#/"]').click()

    def getTotalSmallCart(self, driver):
        total = driver.find_element_by_xpath('//span/label').text
        total = total.replace("(", "").replace(")", "").replace(" Item", "").replace("s", "")
        # return int((driver.find_element_by_css_selector('[class="emptyCart"] [class="roboto-regular"]').text).replace("(", "").replace(")", ""))
        return int(total)

    def getTotalBigCart(self, driver):
        driver.find_element_by_id("shoppingCartLink").click()
        print(self.driver.find_element_by_xpath('//div/section/article/h3').text[15:-1:])
        return self.driver.find_element_by_xpath('//div/section/article/h3').text[15:-1:]


    def clearItemSmall(self, name, color, driver):
        driver.find_element_by_id("shoppingCartLink").click()
        time.sleep(2)
        table = driver.find_element_by_xpath('//tool-tip-cart/div/table/tbody')
        rows = table.find_elements_by_tag_name('tr')
        for row in rows:
            tds = row.find_elements_by_tag_name('td')
            counter = 1
            isThere = 0
            for td in tds:
                if counter == 2:
                    firstIndex = td.text.find("\n")
                    secondIndex = td.text.find("\n", firstIndex + 1)
                    if td.text[:firstIndex:] == name:
                        isThere += 1
                    if td.text[secondIndex + 8::] == color:
                        isThere += 1
                if counter == 3:
                    if isThere == 2:
                        td.find_element_by_css_selector('[icon-x]').click()
                counter += 1




# workbook = load_workbook(filename="DataTests.xlsx", data_only=True)
# sheet = workbook.active
# driver = webdriver.Chrome(executable_path=r"D:\QA Course\Selenium\chromedriver.exe")
# wait = WebDriverWait(driver, 20)
# driver.get("https://www.advantageonlineshopping.com/#/")
# action_chains = ActionChains(driver)
# driver.maximize_window()
# driver.implicitly_wait(1)
# wait.until(EC.visibility_of_element_located((By.ID, "menuUser")))
# time.sleep(5)
# login(driver)
# time.sleep(1)
# clearCart(driver)
# time.sleep(1)
#
# action_chains.move_to_element(driver.find_element_by_id('menuCart')).perform()
# for row in sheet.iter_rows(min_row=2, max_row=4, min_col=2, max_col=4, values_only=True):
#     addProduct(row[0], row[1], row[2], driver)
# time.sleep(2)
#
# #1 1 1 1 1 1 1
# # print("Test 1:", sheet.cell(5, 4) == getTotalSmallCart(driver))
# # time.sleep(1.5)
#
# # 2 2 2 2 2 2
# # for row in sheet.iter_rows(min_row=2, max_row=4, min_col=2, max_col=5, values_only=True):
# #     print(f"is the item: {row[0]} in the cart? ", isInSmallCart(row[0], row[1], row[2], row[3], driver))
#
# # driver.find_element_by_id("shoppingCartLink").click()
# # time.sleep(2)
# # table = driver.find_element_by_xpath('//tool-tip-cart/div/table/tbody')
# # rows = table.find_elements_by_tag_name('tr')
# # isThere = 0
# # for row in rows:
# #     tds = row.find_elements_by_tag_name('td')
# #     counter = 0
# #     for td in tds:
# #         counter += 1
# #         if counter == 2:
# #             print("------------------")
# #             print(type(td), "aaaa")
# #             print(td.text.replace())
# #             print(td.text.replace("\n", "\n iiiiiiiiiiiiiiiiii"))
# #             print("------------------")
# #             # if
# #             attribs = driver.find_elements_by_css_selector('[ng-show="welcome"] [id="toolTipCart"] [class="ng-binding"]')
# #             attribcounter = 1
# #             # print(type(attribs), "  |   ",  attribs[0].text, "  |   ",  attribs[1].text, "  |   ",  attribs[2].text, "  |   ",  attribs[3].text, "  |   ",  attribs[4].text)
# #
# #             # for attrib in attribs:
# #             #     if attribcounter == 1:
# #             #         print("a   ", attrib.text, name)
# #             #         if attrib.text == name:
# #             #             isThere += 1
# #             #     if attribcounter == 2:
# #             #         requa = int(attrib.text.replace("QTY: ", ""))
# #             #         print("a   ", requa, int(quantity))
# #             #         if requa == int(quantity):
# #             #             isThere += 1
# #             #     if attribcounter == 4:
# #             #         print("a   ", attrib.text, color)
# #             #         if attrib.text == color:
# #             #             isThere += 1
# #             #         attribcounter = 1
# #             #     attribcounter += 1
# #
# #         if counter == 3:
# #             print("a   ", td.text.replace(",", "").replace("$", ""))
# #             if float(td.text.replace(",", "").replace("$", "")) == 519.99:
# #                  isThere += 1
# #             counter = 0
# #     if isThere == 4:
# #         print("TRUEEEEEE")
# #     isThere = 0
# # print("FALSEEEEEE")
#
#
#
#
#
#
#
#
#
#
#
# print("begin test \n\n\n\n")
#
#
#
#
#
#
#
#
#
#
#
#
# # driver.find_element_by_id("see_offer_btn").click()
# # time.sleep(3)
# # plus = driver.find_element_by_class_name("plus")
# # plus.click()
# # plus.click()
# # driver.find_element_by_name("save_to_cart").click()
# # driver.find_element_by_xpath("/html/body/div[3]/nav/a[2]").click()
# # driver.find_element_by_id("5").click()
# # driver.find_element_by_name("save_to_cart").click()
# # driver.back()
# # driver.find_element_by_id("1").click()
# # driver.find_element_by_name("save_to_cart").click()
# time.sleep(1)
#
#
#
#
#
# driver.find_element_by_id("shoppingCartLink").click()
# # 2 2 2 2 2 2 2                                  vardi all of 2 is on build right now so ignore it
#
# # table = driver.find_element_by_xpath('//*[@id="shoppingCart"]/table/tbody')
# # rows = table.find_elements_by_tag_name('tr')
# # # name = HP PAVILION 15T TOUCH LAPTOP#----------------------------------------------------------------------------------------------change this to get the requirment from the xl
# # # color = GRAY#----------------------------------------------------------------------------------------------change this to get the requirment from the xl
# # # quantity = 1#----------------------------------------------------------------------------------------------change this to get the requirment from the xl
# # # reqPrice = 519.99#----------------------------------------------------------------------------------------------change this to get the requirment from the xl
# # counter = 0
# # for row in rows:
# #     tds = row.find_elements_by_tag_name('td')
# #     for td in tds:
# #         counter += 1
# #         if counter == 2:
# #             name = "HP PAVILION 15T TOUCH LAPTOP"#-----change this to get the requirment from the xl
# #             # return name == td.text
# #
# #         if counter == 4:
# #             color = "GRAY"  # ----------------------------------------------------------------------------------------------change this to get the requirment from the xl
# #             # return color == td.find_element_by_tag_name('span').get_attribute("title")
# #             print(td.find_element_by_tag_name('span').get_attribute("title"))
# #         if counter == 5:
# #             quantity = 1#----------------------------------------------------------------------------------------------change this to get the requirment from the xl
# #             # return quantity == int(td.text)
# #         if counter == 6:
# #             price = td.find_element_by_tag_name('p').text
# #             currency = price[0]
# #             price = price.replace(",", "").replace("$", "")
# #             price = float(price)
# #             reqPrice = 519.99  # ----------------------------------------------------------------------------------------------change this to get the requirment from the xl
# #             # return reqPrice == price
# #             counter = 0
#
#
#
# # 5 5 5 5 5 5 5 5 5 5 5 5
# table = driver.find_element_by_xpath('//*[@id="shoppingCart"]/table/tbody')
# rows = table.find_elements_by_tag_name('tr')
# for item in sheet.iter_rows(min_row=2, max_row=4, min_col=2, max_col=5, values_only=True):
#     print(isInBigCart(item[0], item[1], item[2], item[3], driver))
#     print()
#
#
#
# time.sleep(8)
# driver.close()




# class inCart:
#     def __init__(self, driver):
#         self.driver

