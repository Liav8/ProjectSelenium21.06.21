from unittest import TestCase
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from usableFunctionalities import *
from Product import *
from CleanCart import *
import time
from openpyxl import *
from main import *
import random


class Tests(TestCase):

    def setUp(self):
        self.page = MainP(webdriver.Chrome(executable_path=r"D:\QA Course\Selenium\chromedriver.exe"))
        # self.page = MainP(webdriver.Firefox(executable_path=r"D:\QA Course\ProjectSelenium21.06.21\SeleniumProject\geckodriver.exe"))
        self.sheet = load_workbook(filename="DataTests.xlsx", data_only=True).active
        wait = WebDriverWait(self.page.driver, 20)
        action_chains = ActionChains(self.page.driver)
        self.page.driver.get("https://www.advantageonlineshopping.com/#/")
        self.page.driver.maximize_window()
        self.page.driver.implicitly_wait(1)
        # wait.until(EC.visibility_of_element_located((By.ID, "menuUser")))
        time.sleep(5)
        login(self.page.driver)
        time.sleep(2)
        clearCart(self.page.driver)
        time.sleep(2)

    def test1(self):
        totalExcel = 0
        for row in self.sheet.iter_rows(min_row=2, max_row=3, min_col=2, max_col=4, values_only=True): #rows 2-3 in the excel are being added
            addProduct(row[0], row[1], row[2], self.page.driver)
            totalExcel += row[2]
        time.sleep(2)
        self.assertEqual(getTotalSmallCart(self.page.driver), totalExcel)

    def test2(self):
        for row in self.sheet.iter_rows(min_row=2, max_row=4, min_col=2, max_col=4, values_only=True): #rows 1-3 in the excel are being added
            addProduct(row[0], row[1], row[2], self.page.driver)
        time.sleep(2)
        for row in self.sheet.iter_rows(min_row=2, max_row=4, min_col=2, max_col=5, values_only=True):
            self.assertTrue(isInSmallCart(row[0], row[1], row[2], row[3], self.page.driver))

    def test3(self):
        for row in self.sheet.iter_rows(min_row=2, max_row=3, min_col=2, max_col=5, values_only=True): #rows 1-3 in the excel are being added
            addProduct(row[0], row[1], row[2], self.page.driver)
        self.page.clearItemSmall("HP PAVILION 15T TOUCH LAPTOP", "RED", self.page.driver)
        self.assertEqual(getTotalSmallCart(self.page.driver), 2)

    def test4(self):
        for row in self.sheet.iter_rows(min_row=2, max_row=2, min_col=2, max_col=5, values_only=True): #row 1 in the excel are being added
            addProduct(row[0], row[1], row[2], self.page.driver)
        self.page.driver.find_element_by_xpath('//*[@id="menuCart"]').click()
        time.sleep(1)
        print(self.page.driver.find_element_by_xpath('//div/section/article/h3').text)
        self.assertEqual("SHOPPING CART", self.page.driver.find_element_by_xpath('//div/section/article/h3').text[:13:])

    def test5(self):
        for row in self.sheet.iter_rows(min_row=2, max_row=4, min_col=2, max_col=5, values_only=True): #rows 1-3 in the excel are being added
            addProduct(row[0], row[1], row[2], self.page.driver)
        time.sleep(2)
        self.page.driver.find_element_by_css_selector('[class="logo"] [href="#/"]').click()
        time.sleep(2)
        self.page.driver.find_element_by_xpath('//*[@id="menuCart"]').click()
        names = getBigCartNames(self.page.driver)
        quantity = getBigCartQuantity(self.page.driver)
        prices = getBigCartPrice(self.page.driver)
        total = 0
        for i in range(0, len(names)):
            print(f"product name is: {names[i]}, product's quantity is: {quantity[i]}, product's total price is: {prices[i]}")
            total += float(prices[i])
        print(self.sheet.cell(5, 5).value, total)
        self.assertEqual(self.sheet.cell(5, 5).value, total)

    def test6(self):
        for row in self.sheet.iter_rows(min_row=2, max_row=3, min_col=2, max_col=5, values_only=True): #rows 1-2 in the excel are being added
            addProduct(row[0], row[1], row[2], self.page.driver)
        self.page.driver.find_element_by_xpath('//*[@id="menuCart"]').click()
        for row in self.sheet.iter_rows(min_row=2, max_row=3, min_col=2, max_col=5, values_only=True):
            updateItem(row[0], row[1], row[1], 4, self.page.driver)
        self.assertEqual(self.page.getTotalBigCart(self.page.driver), 8)

    def test7(self):
        self.page.driver.find_element_by_css_selector('[id="tabletsImg"]').click()
        time.sleep(1)
        self.page.driver.find_element_by_css_selector('[id = "18"]').click()
        time.sleep(1)
        self.page.driver.find_element_by_css_selector('[name="save_to_cart"]').click()
        self.page.driver.back()
        self.page.driver.back()
        print(self.page.driver.find_element_by_css_selector('[id="special_offer_items"] [translate="SPACIAL_OFFER"]').text, "SPECIAL OFFER")
        self.assertEqual(self.page.driver.find_element_by_css_selector('[id="special_offer_items"] [translate="SPACIAL_OFFER"]').text, "SPECIAL OFFER")

    def test8(self):
        self.page.logout(self.page.driver)
        for row in self.sheet.iter_rows(min_row=2, max_row=3, min_col=2, max_col=5, values_only=True): #rows 1-2 in the excel are being added
            addProduct(row[0], row[1], row[2], self.page.driver)
        self.page.driver.find_element_by_xpath('//*[@id="menuCart"]').click()
        self.page.driver.find_element_by_css_selector('[id="checkOutButton"]').click()
        time.sleep(1)
        self.page.driver.find_element_by_css_selector('[id="registration_btnundefined"]').click()
        rand = random
        rnd =str(random.randint(100000, 999999))
        rnd = "Dd" + rnd + ".com"
        self.page.driver.find_element_by_name("usernameRegisterPage").send_keys(rnd)
        self.page.driver.find_element_by_name("passwordRegisterPage").send_keys(rnd)
        self.page.driver.find_element_by_name("confirm_passwordRegisterPage").send_keys(rnd)
        rnd = rnd.replace("d", "@")
        print(rnd)
        self.page.driver.find_element_by_name("emailRegisterPage").send_keys(rnd)
        self.page.driver.find_element_by_css_selector('[name="i_agree"]').click()
        self.page.driver.find_element_by_css_selector('[id="register_btnundefined"]').click()
        time.sleep(15)
#stoped here-----------------------------------------------------------------------------------------------------------stoped here

        self.page.driver.find_element_by_css_selector('[id="next_btn"]').click()
        time.sleep(1)
        self.page.driver.find_element_by_name("safepay_username").send_keys("Sa1234")
        self.page.driver.find_element_by_name("safepay_password").send_keys("Sa1234")
        self.page.driver.find_element_by_css_selector('[id="pay_now_btn_SAFEPAY"]').click()

        #what ever happens next

        self.assertEqual(self.page.getTotalSmallCart(self.page.driver), 0)
        #check if in orders

    def test9(self):
        self.page.logout(self.page.driver)
        for row in self.sheet.iter_rows(min_row=2, max_row=3, min_col=2, max_col=5, values_only=True): #rows 1-2 in the excel are being added
            addProduct(row[0], row[1], row[2], self.page.driver)
        self.page.driver.find_element_by_xpath('//*[@id="menuCart"]').click()
        time.sleep(1)
        self.page.driver.find_element_by_css_selector('[id="checkOutButton"]').click()
        time.sleep(5)
        login(self.page.driver)

        #enter credit card's details

        self.assertEqual(self.page.getTotalSmallCart(self.page.driver), 0)
        # check if in orders

    def test10(self):
        self.page.logout()
        self.assertEqual(self.page.driver.find_element_by_css_selector('[id="menuUserLink"] [data-ng-show="userCookie.response"]'), "")
        login(self.page.driver)
        self.assertEqual(self.page.driver.find_element_by_css_selector('[id="menuUserLink"] [data-ng-show="userCookie.response"]'), "Sa1234")

    def tearDown(self):
        time.sleep(3)
        self.page.driver.close()
